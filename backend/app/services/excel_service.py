from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.models.schemas import StatementData


class ExcelService:
    def build_workbook(self, statements: dict[str, StatementData], output_path: Path) -> None:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        self._build_generic_statement_sheet(workbook.create_sheet(title="Balance Sheet"), statements["balance_sheet"])
        self._build_generic_statement_sheet(workbook.create_sheet(title="P&L"), statements["profit_and_loss"])
        self._build_generic_statement_sheet(workbook.create_sheet(title="Cash Flow"), statements["cash_flow"])

        checks = workbook.create_sheet(title="Validation")
        checks.append(["Statement", "Item", "Year", "Reported Total", "Calculated Total", "Difference", "Status"])
        totals = workbook.create_sheet(title="Totals")
        totals_header = ["Statement", "Item"]
        all_years = []
        for statement in statements.values():
            for year in statement.years:
                if year not in all_years:
                    all_years.append(year)
        totals_header.extend(all_years)
        totals.append(totals_header)
        for statement_name, statement in (
            ("Balance Sheet", statements["balance_sheet"]),
            ("P&L", statements["profit_and_loss"]),
            ("Cash Flow", statements["cash_flow"]),
        ):
            for check in statement.totals_check:
                checks.append(
                    [
                        statement_name,
                        check.item,
                        check.year,
                        check.reported_total,
                        check.calculated_total,
                        check.difference,
                        check.status,
                    ]
                )
            for row in statement.rows:
                if row.item.strip().lower().startswith("total"):
                    totals.append([statement_name, row.item, *(row.values.get(year) for year in all_years)])
        for cell in checks[1]:
            cell.fill = PatternFill("solid", fgColor="2E5B3B")
            cell.font = Font(color="FFFFFF", bold=True)
        for cell in totals[1]:
            cell.fill = PatternFill("solid", fgColor="5A3A18")
            cell.font = Font(color="FFFFFF", bold=True)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)

    def _build_generic_statement_sheet(self, sheet, statement: StatementData) -> None:
        years = statement.years or ["Year 1", "Year 2"]
        header = ["Particulars", *years]
        sheet.append(header)
        for row in statement.rows:
            sheet.append([row.item, *(row.values.get(year) for year in years)])

        self._style_header(sheet, len(years))
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False
        sheet.column_dimensions["A"].width = 52
        for row_index in range(2, sheet.max_row + 1):
            label = str(sheet.cell(row=row_index, column=1).value or "")
            self._style_statement_row(sheet, row_index, label)
        for column_index in range(2, len(years) + 2):
            sheet.column_dimensions[self._column_letter(column_index)].width = 16
            for row_index in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row_index, column=column_index)
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0;(#,##0)'

    def _style_header(self, sheet, year_count: int) -> None:
        fill = PatternFill("solid", fgColor="173C7A")
        font = Font(color="FFFFFF", bold=True)
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center" if cell.column > 1 else "left")


    def _style_statement_row(self, sheet, row_index: int, label: str) -> None:
        lowered = label.strip().lower().rstrip(":")
        thick_side = Side(style="thick", color="000000")
        thin_side = Side(style="thin", color="000000")

        if self._is_major_total(lowered):
            for cell in sheet[row_index]:
                cell.font = Font(bold=True, size=12)
                cell.border = Border(top=thin_side, bottom=thick_side)
        elif self._is_total(lowered):
            for cell in sheet[row_index]:
                cell.font = Font(bold=True)
                cell.border = Border(top=thin_side, bottom=thin_side)
        elif self._is_section_header(lowered):
            sheet.cell(row=row_index, column=1).font = Font(bold=True, size=12)
        elif self._is_subsection_header(lowered):
            sheet.cell(row=row_index, column=1).font = Font(bold=True)
        else:
            sheet.cell(row=row_index, column=1).alignment = Alignment(indent=1)

    def _is_major_total(self, lowered: str) -> bool:
        return lowered.startswith("total assets") or lowered.startswith("total equity and liabilities")

    def _is_total(self, lowered: str) -> bool:
        return lowered.startswith("total") or "net cash flows from" in lowered or "cash generated from / (used in) operations" in lowered

    def _is_section_header(self, lowered: str) -> bool:
        return lowered in {
            "non-current assets",
            "current assets",
            "equity",
            "liabilities",
            "equity and liabilities",
            "non-current liabilities",
            "current liabilities",
            "expenses",
            "income",
            "cash flows from operating activities",
            "cash flows from investing activities",
            "cash flows from financing activities",
            "movements in working capital",
        }

    def _is_subsection_header(self, lowered: str) -> bool:
        return lowered in {
            "fixed assets",
            "financial assets",
            "other current assets",
            "other expenses",
            "cost of goods sold",
            "sg&a expenses",
        }

    def _column_letter(self, index: int) -> str:
        result = ""
        while index:
            index, remainder = divmod(index - 1, 26)
            result = chr(65 + remainder) + result
        return result
